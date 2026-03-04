"""
Sales Bot
=========
Расписание:
- Отчёт по часу:  проверка каждые 5 мин в окна 12:35-13:05, 15:35-16:05, 17:55-18:25, 22:45-23:15
                  отправляет сразу как получит
- День/Неделя/Месяц: письмо приходит в 6-7 утра, бот держит и отправляет в 9:00
  • День     — каждый день
  • Неделя   — только понедельник (после дня, + анализ)
  • Месяц    — 1-е число (анализ с "так и не исправлено:")
             — 15-е или 16-е если пн (анализ с "исправить")
"""

import os, io, json, logging, asyncio
from datetime import datetime, date
import imaplib, email
from email.header import decode_header

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from telegram import Bot

# ── Настройки ──────────────────────────────────────────────────────────────────
TELEGRAM_TOKEN   = os.environ['TELEGRAM_TOKEN']
TELEGRAM_CHAT_ID = os.environ['TELEGRAM_CHAT_ID']
EMAIL_ADDRESS    = os.environ['EMAIL_ADDRESS']
EMAIL_PASSWORD   = os.environ['EMAIL_PASSWORD']
IMAP_SERVER      = os.environ.get('IMAP_SERVER', 'imap.yandex.ru')
SENT_LOG         = 'sent_reports.json'

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

# ── Цвета ──────────────────────────────────────────────────────────────────────
GREEN_FILL  = PatternFill('solid', fgColor='70AD47')
YELLOW_FILL = PatternFill('solid', fgColor='FFD966')
RED_FILL    = PatternFill('solid', fgColor='FF7070')
NO_FILL     = PatternFill('none')

# ── Общая логика заливки ───────────────────────────────────────────────────────
def is_dash(val):
    if val is None: return True
    return str(val).strip() in ('-', '—', '')

def apply_fill(cell, fill):
    cell.fill = fill
    sz   = cell.font.size if cell.font and cell.font.size else 9
    bold = cell.font.bold if cell.font else False
    cell.font = Font(name='Arial', color='000000', size=sz, bold=bold)

def norm(h):
    return '' if h is None else str(h).strip()

def build_thresholds(values, is_zero_red):
    nonzero   = [v for v in values if v != 0]
    has_zeros = any(v == 0 for v in values)

    def thirds(vals):
        if len(vals) < 2: return set(), set(), set(vals)
        sv = sorted(vals); n = len(sv)
        return set(sv[n - n//3:]), set(sv[n//3: n - n//3]), set(sv[:n//3])

    def halves(vals):
        if not vals: return set(), set()
        if len(vals) == 1: return set(vals), set()
        sv = sorted(vals); n = len(sv); mid = n - n//2
        return set(sv[mid:]), set(sv[:mid])

    if is_zero_red and has_zeros:
        green, yellow = halves([v for v in nonzero if v > 0])
        return {'mode': 'halves', 'green': green, 'yellow': yellow}
    else:
        top, mid, bot = thirds(nonzero)
        return {'mode': 'thirds', 'top': top, 'mid': mid, 'bot': bot}

def get_color(val, t, is_zero_red):
    if is_dash(val): return None
    try: v = float(val)
    except: return None
    if v == 0: return RED_FILL
    if not t: return None
    if t['mode'] == 'halves':
        if v < 0: return RED_FILL
        return GREEN_FILL if v in t['green'] else YELLOW_FILL
    else:
        if v in t.get('top', set()): return GREEN_FILL
        if v in t.get('mid', set()): return YELLOW_FILL
        if v in t.get('bot', set()): return RED_FILL
        return YELLOW_FILL

# ── Почасовой отчёт ────────────────────────────────────────────────────────────
H_COLS = [
    'План, %', 'КОП', 'КОП к Нед.', 'КОП к Вчера',
    'ПвЧ', 'ПвЧ к Нед.', 'Штук в чеке',
    'СЧ обувь', 'СЧ Обувь к Нед.', 'СЧ', 'СЧ к Нед.',
    'Доля кожи', 'СЧ Кожа',
    'Расср, чеков', 'Доля\nрассрочки', 'СЧ рассрочка',
    'Конв. об.\nКосм. %', 'Конв.\nстельки, %', 'Аксесс %',
    'ЮИ, %', 'Серебро, %', 'Золото, %', 'Кари Home, %', 'МБТ, %',
    'Косм, %', ' Спорт, % ', 'Сумки,%',
]
H_ZERO_RED = {
    'Расср, чеков', 'Доля\nрассрочки', 'СЧ рассрочка',
    'Конв. об.\nКосм. %', 'Конв.\nстельки, %',
    'Доля кожи', 'СЧ Кожа', 'Аксесс %',
    'ЮИ, %', 'Серебро, %', 'Золото, %', 'Кари Home, %', 'МБТ, %',
    'Косм, %', ' Спорт, % ', 'Сумки,%',
}
H_NO_COLOR = {'ТО к Вчера', 'Трафик к Вчера', 'Доля СБП,%', 'КОП к Вчера'}
H_TO   = 'Ср. ТО к Нед.'
H_TRAF = 'Трафик к Нед.'

def traffic_fill(to_v, traf_v):
    try:
        d = round(float(to_v)*1000)/10 - round(float(traf_v)*1000)/10
    except: return None
    return GREEN_FILL if d >= 0 else (YELLOW_FILL if d >= -1.0 else RED_FILL)

def color_hourly_sheet(ws):
    hrow = None
    for r in ws.iter_rows():
        for c in r:
            if c.value and str(c.value).strip() == 'Магаз':
                hrow = c.row; break
        if hrow: break
    if not hrow: return

    cm = {norm(c.value): c.column for c in ws[hrow] if c.value}
    to_c = cm.get(norm(H_TO)); tr_c = cm.get(norm(H_TRAF))

    rows = []
    for r in ws.iter_rows(min_row=hrow+1, max_row=ws.max_row):
        f = r[0].value
        if f is None: continue
        if str(f).strip() == 'Магаз': break
        try: int(str(f).strip()); rows.append(r)
        except: pass

    hi = {n: cm[norm(n)] for n in H_COLS if norm(n) in cm}
    th = {}
    for n, cidx in hi.items():
        vals = []
        for row in rows:
            v = row[cidx-1].value
            if not is_dash(v):
                try: vals.append(float(v))
                except: pass
        th[n] = build_thresholds(vals, n in H_ZERO_RED)

    for row in rows:
        tf = traffic_fill(
            row[to_c-1].value if to_c else None,
            row[tr_c-1].value if tr_c else None
        )
        for name, cidx in cm.items():
            cell = row[cidx-1]
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == 'FFFFFF00':
                cell.fill = NO_FILL
            if name in {norm(n) for n in H_NO_COLOR}:
                cell.fill = NO_FILL
            elif name in (norm(H_TO), norm(H_TRAF)):
                if tf: apply_fill(cell, tf)
            elif name in {norm(n) for n in H_COLS}:
                fill = get_color(cell.value, th.get(name), name in H_ZERO_RED)
                if fill: apply_fill(cell, fill)
                else: cell.fill = NO_FILL

def process_hourly(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    for sn in wb.sheetnames: color_hourly_sheet(wb[sn])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── Дневной/Недельный/Месячный отчёт ──────────────────────────────────────────
D_COLS = {
    'План %', 'Рост к периоду в ТО', 'ТО LFL',
    'ТО/посетитель', 'ТО/посетитель LFL',
    'КОП', 'Рост к периоду в КОП', 'КОП LFL',
    'КОП обувь', 'Рост к периоду в КОП обувь',
    'Ср. чек', 'Ср. чек LFL', 'Ср. чек обувь', 'Ср. чек обувь LFL',
    'Пар в чеке', 'Пар в чеке LFL',
    'Штук в чеке', 'Штук в чеке к неделе',
    'Ср. пара', 'Ср. пара LFL',
    'Рассрочка %', 'Рассрочка YTY %',
    'Конв. об. косм. %', 'Конв. стельки %', 'Рост к периоду Конв. стельки %',
    'Кожа %', 'Доля РС %', 'Кожа % шт',
    'Аксесс %', 'Аксесс YTY %',
    'Кари Home, %', 'Косм, %', 'Спорт,\xa0%', 'Сумки %', 'Сумки LFL',
    'Качество подбора товара', 'Ср. время сборки ИЗ ч.',
}
D_ZERO_RED = {
    'Рассрочка %', 'Рассрочка YTY %',
    'Конв. об. косм. %', 'Конв. стельки %',
    'Кожа %', 'Доля РС %', 'Кожа % шт',
    'Аксесс %', 'Кари Home, %', 'Косм, %', 'Спорт,\xa0%', 'Сумки %',
    'Ср. время сборки ИЗ ч.',
}
D_NO_COLOR = {
    'Доля СБП %', 'МБТ, %', 'Кидз %', 'Одежда vs Обувь (кидс)',
    'Повт. покуп. LFL', 'Повт. покуп. %', 'Конв. новых клиентов %',
    'Кол-во товаров заказано', 'Доля ИЗ в ТО, %',
}
D_TO   = 'ТО LFL'
D_TRAF = 'Трафик LFL'

def color_day_sheet(ws):
    hrow = None
    for r in ws.iter_rows():
        for c in r:
            if c.value and str(c.value).strip() == 'Магазин':
                hrow = c.row; break
        if hrow: break
    if not hrow: return

    hd = {norm(c.value): c.column for c in ws[hrow] if c.value}
    sc = hd.get('Магазин')
    to_c = hd.get(norm(D_TO)); tr_c = hd.get(norm(D_TRAF))
    if not sc: return

    rows = []
    for r in ws.iter_rows(min_row=hrow+1, max_row=ws.max_row):
        mag = r[sc-1].value
        if mag is None: continue
        if str(mag).strip() == 'Магазин': break
        try: int(str(mag).strip()); rows.append(r)
        except: pass

    hi = {n: hd[norm(n)] for n in D_COLS if norm(n) in hd}
    th = {}
    for n, cidx in hi.items():
        vals = []
        for row in rows:
            v = row[cidx-1].value
            if not is_dash(v):
                try: vals.append(float(v))
                except: pass
        th[n] = build_thresholds(vals, n in D_ZERO_RED)

    for row in rows:
        tf = traffic_fill(
            row[to_c-1].value if to_c else None,
            row[tr_c-1].value if tr_c else None
        )
        for name, cidx in hd.items():
            cell = row[cidx-1]
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == 'FFFFFF00':
                cell.fill = NO_FILL
            if name in {norm(n) for n in D_NO_COLOR}:
                cell.fill = NO_FILL
            elif name in (norm(D_TO), norm(D_TRAF)):
                if tf: apply_fill(cell, tf)
            elif name in {norm(n) for n in D_COLS}:
                fill = get_color(cell.value, th.get(name), name in D_ZERO_RED)
                if fill: apply_fill(cell, fill)
                else: cell.fill = NO_FILL

def process_day(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    for sn in wb.sheetnames: color_day_sheet(wb[sn])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── Извлечение периода из отчёта ───────────────────────────────────────────────
def extract_period(ws, report_type):
    """
    Возвращает строку периода для анализа.
    Row1: 'Период отчета: 25.02.2026. Сформирован ...'
    Row2: 'Период прошедший: 18.02.2026'
    report_type: 'week' | 'month_mid' | 'month_full'
    """
    def get_date(row):
        val = str(ws.cell(row, 1).value or '')
        parts = val.split(':')[-1].strip().split('.')
        if len(parts) >= 3:
            return '.'.join(parts[:3])
        return val.split('.')[0].strip()

    end_date   = get_date(1)   # дата отчёта (конец периода)
    start_date = get_date(2)   # начало периода

    MONTHS_RU = {
        1:'январь',2:'февраль',3:'март',4:'апрель',
        5:'май',6:'июнь',7:'июль',8:'август',
        9:'сентябрь',10:'октябрь',11:'ноябрь',12:'декабрь'
    }

    if report_type == 'week':
        return f"{start_date}–{end_date}"

    elif report_type in ('month_mid', 'month_full'):
        # Пытаемся распарсить месяц из даты отчёта
        try:
            parts = end_date.split('.')
            month = int(parts[1]); year = parts[2]
            month_name = MONTHS_RU.get(month, str(month))
            return f"{month_name} {year} ({start_date}–{end_date})"
        except:
            return f"{start_date}–{end_date}"

    return end_date

# ── Анализ ─────────────────────────────────────────────────────────────────────
def generate_analysis(xlsx_bytes, report_type='week'):
    """
    report_type:
      'week'       → "за неделю DD.MM–DD.MM.YYYY", после ТО: пишет сам % без пометки
      'month_mid'  → "за месяц ...", после ТО: "исправить"
      'month_full' → "за месяц ...", после ТО: "так и не исправлено:"
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes))

    # Находим лист с данными
    ws = None
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for r in sheet.iter_rows():
            for c in r:
                if c.value and str(c.value).strip() == 'Магазин':
                    ws = sheet; break
            if ws: break
        if ws: break
    if not ws: return ''

    hrow = None
    for r in ws.iter_rows():
        for c in r:
            if c.value and str(c.value).strip() == 'Магазин':
                hrow = c.row; break
        if hrow: break
    if not hrow: return ''

    ho = {norm(c.value): c.column for c in ws[hrow] if c.value}
    mc  = ho.get('Магазин');       tc_c = ho.get('ТЦ')
    to_c = ho.get(norm(D_TO));     tr_c = ho.get(norm(D_TRAF))
    tv_c = ho.get('ТО/посетитель LFL')
    kp_c = ho.get('КОП LFL')
    ch_c = ho.get('Ср. чек LFL')
    pa_c = ho.get('Пар в чеке LFL')
    if not mc: return ''

    # Период
    period_str = extract_period(ws, report_type)
    if report_type == 'week':
        header = f"📋 Краткий анализ за неделю {period_str}"
    else:
        header = f"📋 Краткий анализ за месяц {period_str}"

    # Пометка после ТО оборот (жирный текст через Telegram markdown)
    if report_type in ('week', 'month_mid'):
        to_suffix = ' *‼️ ИСПРАВИТЬ*'
    elif report_type == 'month_full':
        to_suffix = ' *❌ НЕ ИСПРАВЛЕНО:*'
    else:
        to_suffix = ''

    def rnd(v): return round(float(v)*1000)/10 if v is not None else None

    stores = []
    for r in ws.iter_rows(min_row=hrow+1, max_row=ws.max_row):
        mag = r[mc-1].value
        if mag is None: continue
        if str(mag).strip() == 'Магазин': break
        try: int(str(mag))
        except: continue

        tc = r[tc_c-1].value if tc_c else ''
        tc_name = str(tc).split(',')[-1].strip() if tc and ',' in str(tc) else str(tc or '')

        to_r = rnd(r[to_c-1].value) if to_c else None
        tr_r = rnd(r[tr_c-1].value) if tr_c else None
        diff = round(to_r - tr_r, 1) if (to_r is not None and tr_r is not None) else None
        worked = diff is not None and diff >= -1.0

        stores.append({
            'mag': mag, 'tc': tc_name,
            'worked': worked, 'diff': diff,
            'topv': rnd(r[tv_c-1].value) if tv_c else None,
            'kop':  rnd(r[kp_c-1].value) if kp_c else None,
            'chek': rnd(r[ch_c-1].value) if ch_c else None,
            'par':  rnd(r[pa_c-1].value) if pa_c else None,
        })

    not_worked  = sorted([s for s in stores if not s['worked']], key=lambda x: x['diff'] or 0)
    worked_list = [s for s in stores if s['worked']]

    lines = [header, "", "🔴 Магазины не отработавшие трафик:"]

    for s in not_worked:
        lines += ["", f"📍 {s['mag']} {s['tc']}"]
        lines.append(f"↘️ ТО {s['diff']:+.0f}%{to_suffix}")
        if s['topv'] is not None and s['topv'] < 0:
            lines.append(f"👤 ТО/посетитель {s['topv']:+.0f}%")
        if s['kop']  is not None and s['kop']  < 0:
            lines.append(f"🛍 КОП {s['kop']:+.0f}%")
        if s['chek'] is not None and s['chek'] < 0:
            lines.append(f"🧾 Ср.чек {s['chek']:+.0f}%")
        if s['par']  is not None and s['par']  < 0:
            lines.append(f"👟 Пар в чеке {s['par']:+.0f}%")

    if worked_list:
        lines += ["", "✅ Отработали трафик:"]
        for s in worked_list:
            lines.append(f"👍 {s['mag']} {s['tc']}")

    lines += ["", "💪 Хороших продаж!"]
    return '\n'.join(lines)

# ── Почта ──────────────────────────────────────────────────────────────────────
def fetch_attachment(keyword):
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        mail.select('INBOX')
        _, data = mail.search(None, 'UNSEEN')
        for mid in reversed(data[0].split()):
            _, msg_data = mail.fetch(mid, '(RFC822)')
            msg = email.message_from_bytes(msg_data[0][1])
            subj = ''
            for part, enc in decode_header(msg.get('Subject', '')):
                subj += part.decode(enc or 'utf-8', errors='replace') if isinstance(part, bytes) else part
            if keyword.lower() not in subj.lower():
                continue
            for part in msg.walk():
                if part.get_content_maintype() == 'multipart': continue
                if not part.get('Content-Disposition'): continue
                fn = ''
                for fp, enc in decode_header(part.get_filename() or ''):
                    fn += fp.decode(enc or 'utf-8', errors='replace') if isinstance(fp, bytes) else fp
                if fn.lower().endswith('.xlsx'):
                    mail.store(mid, '+FLAGS', '\\Seen')
                    mail.logout()
                    return fn, part.get_payload(decode=True)
        mail.logout()
    except Exception as e:
        log.error(f'Email error: {e}')
    return None

# ── Лог отправленных ──────────────────────────────────────────────────────────
def load_sent():
    try:
        with open(SENT_LOG) as f: return json.load(f)
    except: return {}

def already_sent(key): return key in load_sent()

def mark_sent(key):
    d = load_sent(); d[key] = datetime.now().isoformat()
    with open(SENT_LOG, 'w') as f: json.dump(d, f)

# ── Telegram ──────────────────────────────────────────────────────────────────
async def send_file(bot, data, filename, caption):
    await bot.send_document(
        chat_id=TELEGRAM_CHAT_ID,
        document=io.BytesIO(data),
        filename=filename,
        caption=caption,
    )

async def send_text(bot, text):
    await bot.send_message(chat_id=TELEGRAM_CHAT_ID, text=text, parse_mode='Markdown')

# ── Расписание ────────────────────────────────────────────────────────────────
def mid_month_date(year, month):
    """15-е или 16-е если пн."""
    d = date(year, month, 15)
    return date(year, month, 16) if d.weekday() == 0 else d

def in_hourly_window(hour, minute):
    windows = [(12,35,13,5),(15,35,16,5),(17,55,18,25),(22,45,23,15)]
    t = hour*60 + minute
    return any(h1*60+m1 <= t <= h2*60+m2 for h1,m1,h2,m2 in windows)

# ── Главный цикл ──────────────────────────────────────────────────────────────
async def main():
    bot  = Bot(token=TELEGRAM_TOKEN)
    log.info('Бот запущен')

    # Кэш для хранения отчётов полученных до 9:00
    cached = {}  # key -> (filename, raw_bytes)

    while True:
        now     = datetime.now()
        today   = now.date()
        hour    = now.hour
        minute  = now.minute
        weekday = today.weekday()  # 0=пн

        # ══ ПОЧАСОВОЙ — отправляем сразу в окнах ══════════════════════════════
        if in_hourly_window(hour, minute):
            slot = (hour*60 + minute) // 30
            key  = f'hourly_{today}_{slot}'
            if not already_sent(key):
                log.info('Проверяем почасовой...')
                result = fetch_attachment('часу продаж')
                if result:
                    fn, raw = result
                    processed = process_hourly(raw)
                    await send_file(bot, processed, fn, '📊 Отчёт по часу продаж')
                    mark_sent(key)
                    log.info(f'Почасовой отправлен: {fn}')

        # ══ ДЕНЬ/НЕДЕЛЯ/МЕСЯЦ ═════════════════════════════════════════════════

        # Шаг 1: проверяем почту с 6:00 до 8:59 — кэшируем, не отправляем
        if 6 <= hour < 9 and minute < 5:
            cache_key = f'cached_{today}'
            if cache_key not in cached and not already_sent(f'day_{today}'):
                log.info('Проверяем почту (кэш до 9:00)...')
                result = (
                    fetch_attachment('ежедневный отчет') or
                    fetch_attachment('отчет по продажам') or
                    fetch_attachment('продажам за')
                )
                if result:
                    cached[cache_key] = result
                    log.info(f'Отчёт получен и закэширован: {result[0]}')

        # Шаг 2: в 9:00 отправляем всё что накопилось
        if hour == 9 and minute < 5:
            cache_key = f'cached_{today}'
            day_key   = f'day_{today}'

            # Берём из кэша или пробуем получить ещё раз
            if cache_key not in cached and not already_sent(day_key):
                log.info('9:00 — проверяем почту напрямую...')
                result = (
                    fetch_attachment('ежедневный отчет') or
                    fetch_attachment('отчет по продажам') or
                    fetch_attachment('продажам за')
                )
                if result:
                    cached[cache_key] = result

            if cache_key in cached and not already_sent(day_key):
                fn, raw = cached[cache_key]
                log.info(f'9:00 — отправляем отчёты: {fn}')

                # ── ДЕНЬ (каждый день) ────────────────────────────────────────
                processed = process_day(raw)
                await send_file(bot, processed, f'ДЕНЬ_{fn}', '📅 Ежедневный отчёт')
                mark_sent(day_key)
                await asyncio.sleep(2)

                # ── НЕДЕЛЯ (только понедельник) ───────────────────────────────
                if weekday == 0:
                    week_key = f'week_{today}'
                    if not already_sent(week_key):
                        processed_w = process_day(raw)
                        await send_file(bot, processed_w, f'НЕДЕЛЯ_{fn}', '📅 Недельный отчёт')
                        await asyncio.sleep(1)
                        analysis = generate_analysis(raw, report_type='week')
                        if analysis: await send_text(bot, analysis)
                        mark_sent(week_key)
                        await asyncio.sleep(2)

                # ── МЕСЯЦ — 1-е число (полный месяц) ─────────────────────────
                if today.day == 1:
                    m1_key = f'month1_{today.year}_{today.month}'
                    if not already_sent(m1_key):
                        processed_m = process_day(raw)
                        await send_file(bot, processed_m, f'МЕСЯЦ_{fn}', '📅 Месячный отчёт (полный месяц)')
                        await asyncio.sleep(1)
                        analysis = generate_analysis(raw, report_type='month_full')
                        if analysis: await send_text(bot, analysis)
                        mark_sent(m1_key)
                        await asyncio.sleep(2)

                # ── МЕСЯЦ — 15-е (или 16-е если пн) ─────────────────────────
                if today == mid_month_date(today.year, today.month):
                    m15_key = f'month15_{today.year}_{today.month}'
                    if not already_sent(m15_key):
                        processed_m = process_day(raw)
                        await send_file(bot, processed_m, f'МЕСЯЦ15_{fn}', '📅 Месячный отчёт (середина месяца)')
                        await asyncio.sleep(1)
                        analysis = generate_analysis(raw, report_type='month_mid')
                        if analysis: await send_text(bot, analysis)
                        mark_sent(m15_key)

                # Очищаем кэш
                cached.pop(cache_key, None)

        # Пауза: 5 мин в окне почасового, иначе 20 мин
        await asyncio.sleep(5 * 60)  # проверяем каждые 5 минут

if __name__ == '__main__':
    asyncio.run(main())
